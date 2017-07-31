########################################################
# Author: loitd (loitranduc@gmail.com)
# Version: 1.0
# Caution: Don't modify or the program will crashed
# File writter
# openpyxl Version >= 2.3
########################################################
import os, sys
import time
import logging
import openpyxl
from openpyxl.styles import colors, Font, Color, Alignment, Border, Fill, Side
from openpyxl.chart import BarChart, Reference, Series

logging.basicConfig(level=logging.INFO, format='(%(threadName)-10s) %(message)s', )

class FileProcess(object):
    """docstring for FileProcess"""
    def __init__(self, filename):
        super(FileProcess, self).__init__()
        self.filename = filename

    def _checkfile(self):
        if os.path.isfile(self.filename) and os.access(self.filename, os.R_OK):
            return True
        else:
            return False

    def array2file(self, arr):
        logging.info("Begin writting to file")
        try:
            with open(self.filename, 'w+b') as fh:
                for i in arr:
                    line = ','.join(str(j) for j in i)
                    fh.write(line + "\r\n")
            logging.info("Done writing to file.")
        except Exception as e:
            print e

class LogGen(object):
    """docstring for LogGen"""
    def __init__(self):
        super(LogGen, self).__init__()

    def filename_gen(self, pre = "", ext = ".log"):
        # today = time.strftime("%Y-%m-%d %H-%M-%S")
        today = time.strftime("%Y-%m-%d")
        return pre + today + ext

class ExcelProc(FileProcess):
    """docstring for ExcelProc"""
    def __init__(self, filename):
        super(ExcelProc, self).__init__(filename)
        self.wb = None
        self.ws = None

    # get to read only mode
    def getwsread(self, sheetname):
        if self._checkfile():
            wb = openpyxl.load_workbook(self.filename)
            # print("All active sheets: ", wb.get_sheet_names())
            # open the desire sheets
            ws = wb.get_sheet_by_name(sheetname)
            # open active sheet
            # ws = wb.active
            if ws is not None:
                # get the cell value
                self.ws = ws
                return ws
            else:
                logging.warning("Can not get the sheet %s"%sheetname)
        else:
            logging.info("File %s doesn't exits"%self.filename)
        return False

    def readacell(self, cell, sheetname=None):
        # check if user specify sheetname
        # if sheetname is None the ignore getwsread method
        if (sheetname is not None) or (self.ws is None):
            self.getwsread(sheetname)

        cs = self.ws[cell]
        logging.info(cs.value)

    #cells = array of cell name
    def readcells(self, sheetname, cells):
        if (sheetname is not None) or (self.ws is None):
            self.getwsread(sheetname)

        for cell in cells:
            cs = self.ws[cell]
            logging.info(cs.value)

    #cells = array of coordinators
    def readcells2(self, sheetname, cells):
        if (sheetname is not None) or (self.ws is None):
            self.getwsread(sheetname)

        for cell in cells:
            cs = self.ws.cell(row=cell[0], column=cell[1])
            logging.info(cs.value)

    # All writing events need to call this function afterend
    def save(self):
        if self.wb is not None:
            self.wb.save(self.filename)
            logging.info("Done saving file %s"%self.filename)
        else:
            logging.warning("Saving NULL?")

    # just get ws to write
    def getwswrite(self, sheetname):
        try:
            if self._checkfile():
                wb = openpyxl.load_workbook(self.filename)
            else:
                wb = openpyxl.Workbook()

            try:
                # now prepare the sheet
                ws = wb.get_sheet_by_name(sheetname)
            except Exception as e:
                logging.warning("The sheet %s is not available. Now create."%sheetname)
                ws = wb.create_sheet(title=sheetname)

            #return
            self.wb = wb
            self.ws = ws
            logging.info("getwswrite is done.")
            return ws
        except Exception as e:
            logging.warning("Exception while getwswrite")
            print(e)
            return False

    def writeacell(self, sheetname, cell, value):
        if (sheetname is not None) or (self.ws is None):
            self.getwswrite(sheetname)
        # write to cell
        self.ws[cell] = value
        # saving file
        # self.wb.save(self.filename)
        logging.info("Done writing to file %s"%self.filename)

    def writecells(self, cells, values, sheetname=None):
        if len(cells) == len(values):
            if (sheetname is not None) or (self.ws is None):
                # get the ws if need
                self.getwswrite(sheetname)
            # Begin writting
            for i in range(len(cells)):
                self.ws[cells[i]] = values[i]
            # self.wb.save(self.filename)
            logging.info("Done writing to file %s"%self.filename)
        else:
            logging.warning("Cells vs Values don't have equal num of elements. Not permit.")
            return False

    # Merger & Unmerge cells
    # Type = 0 ----> merge
    # Type = 1/others ----> unmerge
    def un_merge(self, sheetname=None, mtype=0, cells='A1:B3'):
        if (sheetname is not None) or (self.ws is None):
            self.getwswrite(sheetname)

        if mtype == 0:
            self.ws.merge_cells(cells)
        else:
            self.ws.unmerge_cells(cells)
        logging.info("Done merge/unmerge cells")

    def addComment(self, cmt, author, cell, sheetname=None):
        if (sheetname is not None) or (self.ws is None):
            self.getwswrite(sheetname)

        comment = openpyxl.comments.Comment(cmt, author)
        # You cannot assign the same Comment object to two different cells.
        self.ws[cell].comment = comment
        logging.info("Done adding comment.")

    def addStyles(self, cell, fontcolor='FF000000', italic=False, bold=True,
                                fontname='Arial', fontsize=14, alignh = 'center',
                                alignv = 'center', wrap_text = True, number_format = 'General',
                                filltype='solid', fill_color='FF00FF00', border='thin',
                                sheetname=None):
        if (sheetname is not None) or (self.ws is None):
            self.getwswrite(sheetname)

        ft = Font(name=fontname, color=fontcolor, size=fontsize, italic=italic, bold=bold)
        alg = Alignment(horizontal=alignh, vertical=alignv, wrap_text=wrap_text)
        # prepare the fontname
        self.ws[cell].font = ft
        self.ws[cell].alignment = alg
        self.ws[cell].number_format = number_format

        if border is not None:
            border_thin = Border(left=Side(border_style=border, color='00000000'),
                                right=Side(border_style=border, color='00000000'),
                                bottom=Side(border_style=border, color='00000000'),
                                top=Side(border_style=border, color='00000000'),
            )
            self.ws[cell].border = border_thin

        logging.info("Add styles done")

    def addFilter(self):
        pass

    def createchart(self, minc = 2, maxc = 3, minr = 3, maxr = 6):
        values = Reference(self.ws, min_col=minc, min_row=minr, max_col=maxc, max_row=maxr)
        cat = Reference(self.ws, min_col=1, min_row=3, max_row=6)
        chart = BarChart()
        chart.title = "Loitd example barchart"
        chart.style = 10
        chart.type = 'col'
        chart.x_axis.title = 'X-Axis Title'
        chart.y_axis.title = 'Y-Axix Title'
        chart.add_data(values, titles_from_data=True)
        chart.set_categories(cat)
        self.ws.add_chart(chart, "G3")


    def writetemplate(self, sheetname=None):
        if (sheetname is not None) or (self.ws is None):
            self.getwswrite(sheetname)

        cells = ('A1',
            'A2', 'B2', 'C2',
            'A3', 'B3', 'C3',
            'A4', 'B4', 'C4',
            'A5', 'B5', 'C5',
            'A6', 'B6', 'C6',
            'A7', 'B7', 'C7',
        )

        values = ('Excel Template Example by Loitd (Loi Tran)',
            'MACHINE', 'IP', 'SESSIONS',
            'DATA06', '1.1.1.1',1000,
            'DATA07', '2.2.2.2',2000,
            'DATA08', '3.3.3.3',3000,
            'DATA09', '4.4.4.4',2000,
            '', 'Total','=SUM(C3:C6)',
        )
        self.writecells(cells, values)
        self.un_merge(cells='A1:C1')
        self.addComment('This is by Loi Tran also. I mean this comment!!!', 'Loi Tran', 'A1')
        self.addStyles(cell='A1', fontcolor='FFFF0000', border='medium')
        self.createchart()
        self.save()


if __name__ == '__main__':
    # fp = FileProcess('abc.info')
    ep = ExcelProc('abc.xlsx')

    #for fast reading
    # ep.getwsread('Loi Tran Demo')

    # ep.readacell(cell='A1')

    # ep.readcells('Loi Tran Demo', ('A1', 'A2', 'A3', 'A4', 'A5'))

    # ep.readcells2('Loi Tran Demo', ((1,1), (1,2), (1,3), (1,4), (1,5)))

    # ep.writeacell('Sheet02', 'A3', '345HGVBNN$%^&^')

    # ep.writecells( ('A3', 'A4', 'B4'), ('345HGVBNN$%^&^', '=SUM(A1,A3)', 1), 'Sheet02')
    ep.getwswrite('Loi Tran Demo')
    # ep.un_merge(sheetname=None)
    # ep.save()

    ep.writetemplate()
